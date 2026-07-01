#!/usr/bin/env python
"""
Script para importar usuários a partir de um arquivo Excel (.xls ou .xlsx).
O usuário importado recebe senha padrão Usu@123.
"""

import os
import re
import sys
import unicodedata
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from werkzeug.security import generate_password_hash

from app import create_app
from app.database import db
from app.models import User

DEFAULT_PASSWORD = "Usu@123"

CPF_DIGITS_RE = re.compile(r"\D+")


def normalize_cpf(value):
    if not value:
        return ""
    return CPF_DIGITS_RE.sub("", str(value))


def make_username(nome):
    nome = normalize_text(nome).strip().lower()
    if not nome:
        return ""
    username = nome.replace(" ", ".")
    username = re.sub(r"[^a-z0-9._-]", "", username)
    username = re.sub(r"\.+", ".", username).strip(".")
    return username or "".join(ch for ch in nome if ch.isalnum())


def ensure_unique_username(base, existing_usernames):
    username = base
    suffix = 1
    while username in existing_usernames or not username:
        username = f"{base}{suffix}"
        suffix += 1
    existing_usernames.add(username)
    return username


def is_header_row(nome, cpf_raw, empresa):
    lower_nome = nome.lower()
    lower_cpf = str(cpf_raw).strip().lower() if cpf_raw is not None else ""
    lower_empresa = empresa.lower()
    if "funcionario" in lower_nome and "cpf" in lower_nome:
        return True
    if lower_nome in {"funcionario", "funcionário", "cpf", "empresa"}:
        return True
    if "cpf" in lower_cpf and ("funcionario" in lower_nome or "empresa" in lower_empresa):
        return True
    return False


def normalize_text(value):
    if not value:
        return ""
    text = str(value).strip()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return text


def parse_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d.%m.%Y", "%d %m %Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def read_excel(filepath):
    filepath_lower = filepath.lower()
    if filepath_lower.endswith(".xls"):
        try:
            import xlrd
        except ImportError:
            raise RuntimeError("xlrd não está instalado. Execute: pip install xlrd")

        workbook = xlrd.open_workbook(filepath)
        sheet = workbook.sheet_by_index(0)
        header_row = None
        for i in range(sheet.nrows):
            row_values = [str(v).strip().lower() for v in sheet.row_values(i)]
            if any("cpf" in v or "funcionário" in v or "funcionario" in v for v in row_values):
                header_row = i
                break
        if header_row is None:
            raise ValueError("Não foi possível encontrar a linha de cabeçalho no .xls")

        headers = [normalize_text(v) for v in sheet.row_values(header_row)]
        rows = []
        for row_idx in range(header_row + 1, sheet.nrows):
            row_values = sheet.row_values(row_idx)
            if any(str(v).strip() for v in row_values):
                rows.append(
                    {
                        headers[i] if i < len(headers) else f"col_{i}": row_values[i]
                        for i in range(len(row_values))
                    }
                )
        return rows

    elif filepath_lower.endswith(".xlsx"):
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl não está instalado. Execute: pip install openpyxl")

        workbook = openpyxl.load_workbook(filepath, data_only=True)
        sheet = workbook.active
        header_row = None
        for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            row_values = [str(v).strip().lower() if v is not None else "" for v in row]
            if any("cpf" in v or "funcionário" in v or "funcionario" in v for v in row_values):
                header_row = i
                break
        if header_row is None:
            raise ValueError("Não foi possível encontrar a linha de cabeçalho no .xlsx")

        headers = [
            normalize_text(v)
            for v in next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
        ]
        rows = []
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if any(str(v).strip() for v in row if v is not None):
                rows.append(
                    {
                        headers[i] if i < len(headers) else f"col_{i}": row[i]
                        for i in range(len(row))
                    }
                )
        return rows

    else:
        raise ValueError("Formato de arquivo não suportado. Use .xls ou .xlsx")


def map_field(headers, options):
    normalized = [h.lower() for h in headers]
    for option in options:
        for i, header in enumerate(normalized):
            if option in header:
                return i
    return None


def import_users(filepath):
    rows = read_excel(filepath)
    if not rows:
        print("Nenhuma linha encontrada no arquivo.")
        return

    first_row = rows[0]
    headers = list(first_row.keys())
    cpf_idx = map_field(headers, ["cpf"])
    nome_idx = map_field(headers, ["funcionário", "funcionario", "nome"])
    tipo_idx = map_field(
        headers, ["tipo vínculo", "tipo vinculo", "tipo contrato", "vínculo", "vinculo"]
    )
    funcao_idx = map_field(headers, ["função", "funcao", "cargo", "role"])
    admissao_idx = map_field(
        headers, ["admissão", "admissao", "data_admissao", "data admissao", "hire_date"]
    )
    empresa_idx = map_field(headers, ["empresa", "company"])

    if cpf_idx is None or nome_idx is None or empresa_idx is None:
        raise ValueError("O arquivo deve conter pelo menos as colunas CPF, Funcionário e Empresa.")

    app = create_app()
    with app.app_context():
        imported = 0
        updated = 0
        skipped = 0
        errors = []

        existing_users = User.query.all()
        existing_usernames = {user.username for user in existing_users if user.username}
        existing_by_cpf = {
            normalize_cpf(user.cpf): user for user in existing_users if normalize_cpf(user.cpf)
        }

        for row_number, row in enumerate(rows, start=1):
            cpf_raw = row.get(headers[cpf_idx], "")
            cpf = normalize_text(cpf_raw)
            cpf_norm = normalize_cpf(cpf_raw)
            nome = normalize_text(row.get(headers[nome_idx], ""))
            tipo_vinculo = (
                normalize_text(row.get(headers[tipo_idx], "")) if tipo_idx is not None else "CLT"
            )
            funcao = (
                normalize_text(row.get(headers[funcao_idx], "")) if funcao_idx is not None else ""
            )
            admissao_raw = row.get(headers[admissao_idx]) if admissao_idx is not None else None
            empresa = normalize_text(row.get(headers[empresa_idx], ""))

            if is_header_row(nome, cpf_raw, empresa):
                skipped += 1
                errors.append(f"Linha {row_number + 1}: linha de cabeçalho/resumo ignorada")
                continue

            if not cpf or not nome or not empresa:
                skipped += 1
                errors.append(f"Linha {row_number + 1}: campos obrigatórios ausentes")
                continue

            username_candidate = make_username(nome)
            if not username_candidate:
                skipped += 1
                errors.append(f"Linha {row_number + 1}: nome inválido para username")
                continue

            user = None
            if cpf_norm and cpf_norm in existing_by_cpf:
                user = existing_by_cpf[cpf_norm]
            elif username_candidate in existing_usernames:
                user = User.query.filter_by(username=username_candidate).first()

            if user is None:
                username = username_candidate
                if username in existing_usernames:
                    username = ensure_unique_username(username_candidate, existing_usernames)
                senha_hash = generate_password_hash(DEFAULT_PASSWORD)
                user = User(
                    username=username,
                    password=senha_hash,
                    role="usuario",
                    tipo_contrato="CLT",
                    cpf=cpf,
                    cargo=funcao,
                    empresa=empresa,
                    data_admissao=parse_date(admissao_raw),
                )
                db.session.add(user)
                imported += 1
                existing_usernames.add(username)
                if cpf_norm:
                    existing_by_cpf[cpf_norm] = user
                print(f"Criado: {username} / {cpf} / {empresa}")
            else:
                user.cpf = cpf
                user.cargo = funcao
                user.empresa = empresa
                user.tipo_contrato = (
                    "PJ"
                    if (
                        "pj" in tipo_vinculo.lower()
                        or "pessoa juridica" in tipo_vinculo.lower()
                        or "pessoa jurídica" in tipo_vinculo.lower()
                    )
                    else "CLT"
                )
                user.data_admissao = parse_date(admissao_raw)
                user.password = generate_password_hash(DEFAULT_PASSWORD)
                updated += 1
                print(f"Atualizado: {user.username} / {cpf} / {empresa}")

        try:
            db.session.commit()
            print(
                f"\nImportação finalizada: {imported} usuário(s) criados, {updated} usuário(s) atualizados, {skipped} linha(s) pulada(s)."
            )
            if errors:
                print("Erros:")
                for err in errors:
                    print(" -", err)
                print(
                    "\nTodos os usuários importados ou atualizados receberão a senha fixa: Usu@123"
                )
        except Exception as exc:
            db.session.rollback()
            raise RuntimeError(f"Erro ao salvar no banco: {exc}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Uso: python scripts/import_users.py <arquivo.xls ou arquivo.xlsx>")
        sys.exit(1)

    filepath = sys.argv[1]
    if not os.path.exists(filepath):
        print(f"Arquivo não encontrado: {filepath}")
        sys.exit(1)

    import_users(filepath)
